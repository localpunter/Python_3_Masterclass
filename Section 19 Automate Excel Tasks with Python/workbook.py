from openpyxl import load_workbook, workbook, worksheet

import openpyxl

# WORKING WITH CELL AND CELL INFO
# workbook = load_workbook(
#     "C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\Employees.xlsx")

# sheet = workbook["Skills"]

# print(workbook.properties)
# print(workbook.sheetnames)
# print(workbook.active)
# print(sheet)

# workbook.create_sheet("TestSheet")
# workbook.create_sheet("TestSheet1")
# workbook.save("C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\Employees.xlsx")

# sheet = workbook["TestSheet"]
# workbook.remove(sheet)
# del workbook["TestSheet1"]

# workbook.save("C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\Employees.xlsx")
# workbook.close()

# sheet = workbook["EmployeeData"]

# Setting the title of a sheet or renaming the sheet
# sheet.title = 'Employees'

# print(sheet.title)  # prints the title of the sheet
# print(sheet.active_cell)  # prints the first active cell in the sheet (usually A1)
# # print(dir(sheet))  # prints the available methods and attributes when working with sheets
# print(sheet.dimensions)  # prints the cell range where data is entered
# print("This is sheet_format", sheet.sheet_format)  # prints the basic parameters of the sheet, row height, col width etc
# print("This is sheet_dimensions", sheet.sheet_properties)  # prints more parameters of the sheet
# print("This is sheet_state", sheet.sheet_state)  # prints whether the sheet is hidden or visible
# print("This is max number of rows in use", sheet.max_row)  # prints the max number of rows in use
# print("This is max number of columns in use", sheet.max_column)  # prints the max number of columns in use

# This iterates over each row and prints the rows in the form of a tuple
# for i in sheet.values:
#     print(i)

# sheet = workbook["EmployeeData"]

# # Getting the value of a cell in a sheet, by the cell coordinates
# print("This is currently the value of B10 & C10:", sheet["B10"].value, sheet["C10"].value)
# # or
# print(sheet.cell(row = 5, column = 3).value)

# cell = sheet["B9"]  # Selecting a cell in the sheet

# print("This returns the row reference:", cell.row)
# print("This returns the column reference:", cell.column)
# print("This returns the column and row reference:", cell.coordinate)
# print("This returns the data type of the cell:", cell.data_type)
# TYPE_STRING = 's'
# TYPE_FORMULA = 'f'
# TYPE_NUMERIC = 'n'
# TYPE_BOOL = 'b'
# TYPE_NULL = 'n'
# TYPE_INLINE = 'inlineStr'
# TYPE_ERROR = 'e'
# TYPE_FORMULA_CACHE_STRING = 'str'
# print("This returns the encoding format of the cell:", cell.encoding)

# cell = sheet["B10"]
# cell.value = "Drakey"
# workbook.save("C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\Employees.xlsx")
# print("This is the new value of B10 & C10:", sheet["B10"].value, sheet["C10"].value)

# print("This returns the parent attribute of the cell:", cell.parent)


# WORKING WITH CELL STYLES

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

# print(dir(openpyxl.styles))

workbook = load_workbook(
    "C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\Employees.xlsx")

sheet = workbook["EmployeeData"]
cell = sheet["B8"]

font = Font(color = colors.GREEN, bold = True, italic = True)

cell.font = font

fill = PatternFill(fill_type = "solid", bgColor = "D7F442")

cell.fill = fill

border = Border(
    left = Side(border_style = "double", color = "45f442"),
    right = Side(border_style = "double", color = "45f442"),
    top = Side(border_style = "double", color = "45f442"),
    bottom = Side(border_style = "double", color = "45f442")
    )

cell.border = border

align = Alignment(horizontal = "left")

cell.alignment = align

workbook.save("C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\Employees.xlsx")
