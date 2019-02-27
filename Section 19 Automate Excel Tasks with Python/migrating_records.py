# Importing the necessary modules
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.worksheet.table import Table, TableStyleInfo, TABLESTYLES

# Opening the text file for reading
text_file = open(
    "C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\employees.txt")

# Creating an empty list for storing the records as a list of lists
records = []

# Making sure we are reading the file from the very beginning
text_file.seek(0)

# Splitting each line in the file by the ";" delimiter and appending each list generated by readlines() to the new list, records
for record in text_file.readlines():
    records.append(record.rstrip("\n").split(";"))

# Printing the 'records' list
# print(records)

# Creating a new workbook object, by initializing the Workbook() class
workbook = Workbook()

# Setting the path to (location of) the new Excel workbook
# Note: we should escape the "\" characters inside the path using a "\", in order to avoid any conflicts with Python's special characters, like \n (new line ) or \t (tab).
file_path = "C:\\Users\\local_4nmktoc\\Documents\\Coding\\Python\\Python_3_MasterClass\\Section 19 Automate Excel Tasks with Python\\MyCompanyStaff.xlsx"

# Saving the workbook
workbook.save(file_path)

# Printing out the existing sheets (the default sheet is called 'Sheet')
# print(workbook.sheetnames)

# Renaming the default sheet to 'Employees'
sheet = workbook["Sheet"]
sheet.title = "Employees"

# Referencing the sheet
sheet = workbook['Employees']

# Populating the sheet with the rows from the text file
# Note that the table header will be created from the first row in the text file
for row in records:
    sheet.append(row)

# Creating a table inside the sheet
table = Table(displayName="Table", ref="A1:G11")

# #Defining a style for the table (default style name, row/column stripes)
# #Choose your table style from the default styles of openpyxl
# #Just type in print(openpyxl.worksheet.table.TABLESTYLES) in Python
style = TableStyleInfo(name="TableStyleMedium8",
                       showRowStripes=True, showColumnStripes=False)

# Applying the style to the table
table.tableStyleInfo = style

# Adding the newly created table to the sheet
sheet.add_table(table)

# Defining the font (red, bold, italic) for salary > 55000
font = Font(color=colors.RED, bold=True, italic=True)

# Applying the font settings to the cells that meet the condition salary > 55000
for cell_no in range(2, 12):
    if int(sheet["G%s" % (cell_no)].value) > 55000:
        sheet["G%s" % (cell_no)].font = font

# Saving the changes made to the workbook
workbook.save(file_path)

# #Closing the text file
text_file.close()

# #Closing the workbook
workbook.close()
